import xlrd
xlsx=xlrd.open_workbook(r"C:\Users\Administrator\Desktop\南山区_自主上报_未落块地址_截至0226 125952.xlsx")
table=xlsx.sheet_by_index(1)
table=xlsx.sheet_by_name('Sheet1')
# 单元格的值-从0开始
print(table.cell_value(5,2))


import xlwt
new_workbook=xlwt.Workbook()
worksheet=new_workbook.add_sheet('new_sheet')
worksheet.write(0,0,'new_sheet')
new_workbook.save(r"C:\Users\Administrator\Desktop\xlwt_test.xls")

import xlutils.copy import copy
import xlrd
import xlwt
tem_excel=xlrd.open_workbook('',formatting_info=True)
tem_sheet=tem_excel.sheet_by_index(0)

new_excel=copy(tem_excel)
new_sheet=new_excel.get_sheet(0)

# 初始化一个样式
style=xlwt.XFStyle()
font=style.Font()
font.name='微软雅黑'
font.bold=True
font.height=18*20
# 字体添加进去
style.font=font

# 设置边框
borders=xlwt.Borders()
borders.top=xlwt.Borders.THIN
borders.bottom=xlwt.Borders.THIN
borders.left=xlwt.Borders.THIN
borders.right-xlwt.Borders.THIN

# 设置对齐
alignment=xlwt.Alignment()
alignment.horz=xlwt.Alignment.HORZ_CENTER
alignment.vert=xlwt.Alignment.VERT_BOTTOM





import xlwt
import xlrd
xlsx=xlrd.open_workbook(r"C:\Users\Administrator\Desktop\南山区_自主上报_未落块地址_截至0226 125952.xlsx")
table=xlsx.sheet_by_index(1)
table=xlsx.sheet_by_name('Sheet1')
# 单元格的值-从0开始
all_data=[]
for n in range(1,table.nrows):
    print(table.cell(n,1).value)
    all_data.append(table.cell(n,1).value)

tem_excel=xlrd.open_workbook('D:/模板.xls',formatting_info=True)
tem_sheet=tem_excel.sheet_by_name('Sheet1')

new_excel=copy(tem_excel)
new_sheet=new_excel.get_sheet(0)

style=xlwt.XFStyle()

font=xlwt.Font()
font.name='微软雅黑'
font.bold=True
font.height=360
style.font=font

borders=xlwt.Borders()
borders.top=xlwt.Borders.THIN
borders.bottom=xlwt.Borders.THIN
borders.left=xlwt.Borders.THIN
borders.right=xlwt.Borders.THIN
style.borders=borders

alignment=xlwt.Alignment()
alignment.horz=xlwt.Alignment.HORZ_CENTER
alignment.vert=xlwt.Alignment.VERT_BOTTOM
style.alignment=alignment

new_sheet.write(2,1,5,style)

new_excel.save(r'')




import xlwt
workbook=xlwt.Workbook()
sheet0=workbook.add_sheet('sheet0')
for i in range(0,500):
    sheet0.write(0,i,1)

# xlsxwriter不支持带格式的文件
import xlsxwriter as xw
workbook=xw.Workbook(r'd:/text.xlsx')
sheet0=workbook.add_worksheet('sheet0')
for i in range(0,300):
    sheet0.write(0,i,i)
workbook.close()

# openpyxl性能不稳定
import openpyxl
# 载入已有的工作簿
workbook=openpyxl.load_workbook(r'统计模板.xlsx')
sheet0=workbook['sheet1']
sheet0['b3']='5'
workbook.save(r'测试写数据.xlsx')





import os
import xlwt
file_dir='d:/'
os.listdir(file_dir)
new_workbook=xlwt.Workbook()
worksheet=new_workbook.add_sheet('new_sheet')
n=0
for i in os.listdir(file_dir):
    worksheet.write(n,0,i)
    n+=1
new_workbook.save(r'')









# 发送邮件

import smtplib
# 加密邮件内容，防止中途被截获
from smtplib import SMTP_SSL
# 构造邮件正文
from email.mime.text import MIMEText
# 把邮件各个部分装在一起
from email.mime.multipart import MIMEMultipart
# 邮件头：收件人，标题等
from email.header import  import Header
host_server='smtp.sina.com' # sina邮箱的smtp服务器
sender_sina='pythonauto@sina.com' # sender_sina为发件人的邮箱
pwd='python1234'

sender_sina_mail='python@sina.com' #发件人的邮箱
receiver='pythonto@sina.com'# 收件人的邮箱

mail_title='python办公自动化'# 邮件标题
mail_content='你好，这是使用python登录sina邮箱发送邮件的测试'# 邮件正文

msg=MIMEMultipart()# 初始化邮件主体
msg['Subject']=Header(mail_title,'utf-8')
msg['From']=sender_sina_mail
msg['To']=Header('测试邮箱','utf-8')
msg.attach(MIMEText(mail_content,'plain','utf-8'))# 邮件正文内容 plain无格式邮件

smtp=SMTP_SSL(host_server)# ssl登录
smtp.login(sender_sina_mail,receiver,msg.as_string())
smtp.quit()





import openpyxl
wb=openpyxl.Workbook()
ws=wb.active
ws['a1']=23
ws.append([1,2,3])
ws.append([4,5,6])
import datetime
ws['a4']=datetime.datetime.now()
wb.save(r"C:\Users\Administrator\Desktop\test.xlsx")

def save_to_excel(result):
    wb=openpyxl.Workbook()
    ws=wb.active
    ws['a1']='电影名称'
    ws['b1']='评分'
    ws['c1']='资料'
    for each in result:
        ws.append(each)
    wb.save(r"C:\Users\Administrator\Desktop\test.xlsx")
if __name__ == '__main__':
    result=[('1','a','aa'),('2','b','bb')]
    save_to_excel(result)


	import openpyxl
wb=openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\test.xlsx")
wb.get_active_sheet()
# 获取已有工作表名称
sheets_names=wb.get_sheet_names()
sheets_names=wb.sheetnames
ws=wb.get_sheet_by_name('Sheet1')
# 创建工作表
wb.create_sheet(index=1,title='sht1')
wb.sheetnames
# 删除工作表
wb.remove_sheet(wb.get_sheet_by_name('sht1'))
wb.sheetnames
# 定位单元格 row column coordinate
c=ws['a2']
c.row
c.column
c.coordinate
# 单元格的值
c.value
# 单元格偏移
d=c.offset(2,0)
d.coordinate
# 获取第N列的列名称
openpyxl.cell.cell.get_column_letter(444)
#从列名称获取第N列
openpyxl.cell.cell.column_index_from_string('ggg')


# 访问多个单元格
for each_movie in ws['a2':'b20']:
    for each_cell in each_movie:
        print(each_cell.value,end='')
    print('\n')
# 先行后列打印

# 打印第一列
for each_row in ws.row:
    print(each_row[0].value)
for each_row in ws.iter_row(min_row=2,min_col=1,max_row=4,max_col=8):
    print(each_row[0].value)

# 拷贝工作表
new=wb.copy_worksheet(ws)
wb.save(r"C:\Users\Administrator\Desktop\test.xlsx")

# 个性化标签栏
import openpyxl
wb=openpyxl.Workbook()
ws1=wb.create_sheet(title='sht1')
ws2=wb.create_sheet(title='sht2')
ws3=wb.create_sheet(title='sht3')
ws1.sheet_properties.tabColor='00FF0000'
wb.save(r"C:\Users\Administrator\Desktop\demo.xlsx")

# 调整行高和列宽
# row_dimensions[2].height 第二行高度
# column_dimensions['c'].width 第C列宽度
ws2.row_dimensions[2].heigth=100
ws2.column_dimensions['c'].width=100
wb.save(r"C:\Users\Administrator\Desktop\demo.xlsx")


# 合并和拆分单元格
ws1.merge_cells('a1:c3')
ws1['a1']=3
wb.save(r"C:\Users\Administrator\Desktop\demo.xlsx")

ws1.unmerge_cells('a1:c3')
ws1['a1']=3
wb.save(r"C:\Users\Administrator\Desktop\demo.xlsx")

# 冻结
import openpyxl
wb=openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\demo.xlsx")
ws=wb.active
ws.freeze_panes='b2'
wb.save(r"C:\Users\Administrator\Desktop\demo.xlsx")
